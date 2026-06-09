"""
Products Pipeline API (port 8006)
파일 업로드 → 상품 일괄 등록.
파일 업로드는 두 단계로 처리된다:
  1. POST /products/register/upload  → job_id 즉시 반환
  2. GET  /products/jobs/{job_id}/stream  → SSE 진행상황 스트리밍 (재연결 지원)
"""

import asyncio
import csv
import io
import json
from typing import Any, Dict, List

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse

from ..config.settings import settings
from ..core.auth import UserContext
from ..core.logging import get_logger
from .deps import require_admin_from_headers
from .upload_jobs import UploadJob, append_event, create_job, get_events_after, get_job

logger = get_logger("products_pipeline")

router = APIRouter(prefix="/api/pipeline", tags=["Products Pipeline"])



# ──────────────────────────────────────────────────────
# 파일 파싱
# ──────────────────────────────────────────────────────

_FORMULA_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_formula(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    # 임베디드 newline은 CSV export 시 셀 분리를 유발하므로 공백으로 치환
    value = value.replace("\n", " ").replace("\r", " ")
    if value.startswith(_FORMULA_CHARS):
        return "'" + value
    return value


def _parse_image_field(value: Any) -> list:
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return []
        if value.startswith("["):
            try:
                return json.loads(value)
            except json.JSONDecodeError:
                pass
        return [u.strip() for u in value.split(",") if u.strip()]
    return []


def _normalize_record(record: Dict[str, Any]) -> Dict[str, Any]:
    IMAGE_KEYS = {"상품상세_이미지", "상품이미지", "image_urls"}
    result = {}
    for k, v in record.items():
        if k in IMAGE_KEYS:
            result[k] = _parse_image_field(v)
        else:
            result[k] = _sanitize_formula(v)
    return result


def _parse_file_to_records(filename: str, content: bytes) -> List[Dict[str, Any]]:
    name_lower = filename.lower()

    if name_lower.endswith(".jsonl"):
        records = []
        skipped = 0
        for line_no, raw in enumerate(content.decode("utf-8").splitlines(), start=1):
            line = raw.strip()
            if not line:
                continue
            try:
                records.append(_normalize_record(json.loads(line)))
            except json.JSONDecodeError as exc:
                skipped += 1
                logger.warning(
                    "jsonl_parse_error",
                    filename=filename,
                    line_no=line_no,
                    preview=line[:50],
                    error_type=type(exc).__name__,
                )
        if skipped:
            logger.warning(
                "jsonl_lines_skipped",
                filename=filename,
                skipped=skipped,
                parsed=len(records),
            )
        return records

    if name_lower.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = content.decode("euc-kr")
            except UnicodeDecodeError:
                raise ValueError("지원하지 않는 인코딩입니다. UTF-8 또는 EUC-KR을 사용해주세요.")
        text_io = io.StringIO(text)
        reader = csv.DictReader(text_io)
        return [_normalize_record(dict(row)) for row in reader]

    if name_lower.endswith(".xlsx"):
        import zipfile

        try:
            import openpyxl
        except ImportError as e:
            raise ValueError("XLSX 파일 처리를 지원하지 않습니다. 관리자에게 문의해주세요.") from e
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except (openpyxl.utils.exceptions.InvalidFileException, zipfile.BadZipFile) as e:
            raise ValueError("손상되었거나 지원하지 않는 XLSX 파일입니다.") from e
        ws = wb.active
        if ws is None:
            raise ValueError("XLSX 파일에 활성 시트가 없습니다.")
        headers = [cell.value for cell in ws[1]]
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {k: v for k, v in zip(headers, row) if k is not None}
            records.append(_normalize_record(record))
        return records

    raise ValueError(f"지원하지 않는 파일 형식: {filename} (JSONL, CSV, XLSX만 허용)")


# ──────────────────────────────────────────────────────
# 백그라운드 워커
# ──────────────────────────────────────────────────────

async def _run_product_job(
    job: UploadJob,
    records: List[Dict[str, Any]],
    service: Any,
    creator_user_id: str,
) -> None:
    """파일 내 모든 레코드를 처리하고 결과를 job 이벤트 버퍼에 기록한다."""
    queue: asyncio.Queue = asyncio.Queue()
    semaphore = asyncio.Semaphore(settings.upload_product_concurrency)
    chunk_size = settings.upload_product_concurrency * 4
    active_tasks: list[asyncio.Task] = []

    async def process_and_enqueue(record: dict) -> None:
        try:
            async with semaphore:
                result = await service.register_product(record, user_id=creator_user_id)
        except Exception:
            logger.error("register_product_failed", product_name=record.get("상품명", ""), exc_info=True)
            result = {
                "success": False,
                "product_name": record.get("상품명", ""),
                "error": "상품 등록 중 오류가 발생했습니다.",
            }
        await queue.put(result)

    try:
        succeeded = 0
        failed = 0
        for chunk_start in range(0, len(records), chunk_size):
            chunk = records[chunk_start : chunk_start + chunk_size]
            active_tasks = [
                asyncio.create_task(process_and_enqueue(r))
                for r in chunk
            ]
            for _ in range(len(chunk)):
                result = await queue.get()
                if result.get("success"):
                    succeeded += 1
                else:
                    failed += 1
                await append_event(job, {
                    "type": "progress",
                    "current": succeeded + failed,
                    "total": job.total,
                    "name": result.get("product_name"),
                    "success": result.get("success", False),
                    "product_id": result.get("product_id"),
                    "error": result.get("error"),
                })
            await asyncio.gather(*active_tasks, return_exceptions=True)
            active_tasks = []

        logger.info(
            "register_products_job_completed",
            job_id=job.job_id,
            total=job.total,
            succeeded=succeeded,
            failed=failed,
        )
        await append_event(job, {
            "type": "done",
            "total": job.total,
            "succeeded": succeeded,
            "failed": failed,
        })
    except Exception:
        logger.error("register_products_job_failed", job_id=job.job_id, exc_info=True)
        await append_event(job, {
            "type": "error",
            "detail": "상품 등록 중 오류가 발생했습니다.",
        })
    finally:
        for task in active_tasks:
            task.cancel()
        if active_tasks:
            await asyncio.gather(*active_tasks, return_exceptions=True)


async def _guarded_run_product_job(
    job: UploadJob,
    records: list[dict],
    service: Any,
    creator_user_id: str,
) -> None:
    try:
        await asyncio.wait_for(
            _run_product_job(job, records, service, creator_user_id),
            timeout=settings.upload_job_max_seconds,
        )
    except asyncio.TimeoutError:
        logger.error(
            "product_job_timed_out",
            job_id=job.job_id,
            timeout=settings.upload_job_max_seconds,
        )
        try:
            await append_event(job, {"type": "error", "detail": "상품 등록 시간이 초과됐습니다."})
        except Exception:
            logger.error("product_job_status_update_failed", job_id=job.job_id)
    except Exception:
        logger.error("product_job_outer_crashed", job_id=job.job_id, exc_info=True)
        try:
            await append_event(job, {"type": "error", "detail": "상품 등록 중 오류가 발생했습니다."})
        except Exception:
            logger.error("product_job_status_update_failed", job_id=job.job_id)


# ──────────────────────────────────────────────────────
# SSE 스트림 제너레이터
# ──────────────────────────────────────────────────────

async def _stream_job_events(job: UploadJob):
    """
    재연결을 지원하는 SSE 이벤트 스트림 (PostgreSQL polling).
    Phase A: 기존 이벤트 즉시 replay → Phase B: 0.5초 간격으로 새 이벤트 polling.
    sse_keepalive_timeout초마다 keepalive 코멘트를 전송해 프록시 타임아웃을 방지한다.
    sse_stream_max_seconds 경과 시 error 이벤트를 전송하고 종료해 orphan 연결을 방지한다.
    """
    last_id = 0
    keepalive_ticks = 0
    keepalive_interval = max(1, int(settings.sse_keepalive_timeout / 0.5))
    first = True
    deadline = asyncio.get_event_loop().time() + settings.sse_stream_max_seconds

    while True:
        if asyncio.get_event_loop().time() >= deadline:
            yield f"data: {json.dumps({'type': 'error', 'detail': '스트림 최대 시간이 초과됐습니다.'}, ensure_ascii=False)}\n\n"
            return
        if not first:
            await asyncio.sleep(0.5)
        first = False

        event_rows = await get_events_after(job.job_id, after_id=last_id)
        if event_rows:
            for event_id, event in event_rows:
                last_id = event_id
                yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
                if event.get("type") in ("done", "error"):
                    return
            keepalive_ticks = 0
        else:
            keepalive_ticks += 1
            if keepalive_ticks >= keepalive_interval:
                yield ": keepalive\n\n"
                keepalive_ticks = 0


# ──────────────────────────────────────────────────────
# 엔드포인트
# ──────────────────────────────────────────────────────

@router.post("/products/register/upload")
async def upload_products_file(
    file: UploadFile = File(...),
    request: Request = None,
    current_user: UserContext = Depends(require_admin_from_headers),
) -> Dict[str, Any]:
    """
    JSONL / CSV / XLSX 파일을 업로드하고 백그라운드 처리를 시작한다.
    즉시 job_id를 반환하며, 진행상황은 /products/jobs/{job_id}/stream 으로 수신한다.
    """
    try:
        content = await asyncio.wait_for(
            file.read(settings.max_upload_bytes + 1),
            timeout=settings.upload_file_read_timeout,
        )
    except asyncio.TimeoutError:
        raise HTTPException(status_code=408, detail="파일 읽기 시간이 초과되었습니다.")

    if len(content) > settings.max_upload_bytes:
        raise HTTPException(status_code=413, detail="파일 크기는 50MB를 초과할 수 없습니다.")

    try:
        records = await asyncio.wait_for(
            asyncio.to_thread(_parse_file_to_records, file.filename or "", content),
            timeout=settings.upload_file_parse_timeout,
        )
    except asyncio.TimeoutError:
        raise HTTPException(status_code=408, detail="파일 파싱 시간이 초과되었습니다.")
    except ValueError:
        logger.warning("file_parse_failed", filename=file.filename)
        raise HTTPException(status_code=422, detail="파일 형식이 올바르지 않습니다.")

    if not records:
        raise HTTPException(status_code=422, detail="파일에서 유효한 레코드를 찾을 수 없습니다.")

    if len(records) > settings.max_records_per_upload:
        raise HTTPException(
            status_code=422,
            detail=f"레코드 수가 최대 허용치({settings.max_records_per_upload}개)를 초과합니다. 현재: {len(records)}개",
        )

    service = request.app.state.registration
    job = await create_job("product", len(records), creator_user_id=current_user.user_id)
    if job is None:
        raise HTTPException(
            status_code=409,
            detail="활성 업로드 작업이 너무 많습니다. 기존 작업이 완료된 후 다시 시도하세요.",
        )

    asyncio.create_task(
        _guarded_run_product_job(job, records, service, current_user.user_id)
    )

    return {"job_id": job.job_id, "total": job.total}


@router.get("/products/jobs/{job_id}/stream")
async def stream_product_job(
    job_id: str,
    request: Request = None,
    current_user: UserContext = Depends(require_admin_from_headers),
) -> StreamingResponse:
    """job_id에 해당하는 상품 등록 진행상황을 SSE로 스트리밍한다. 재연결 시 처음부터 replay된다."""
    job = await get_job(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다.")
    if job.job_type != "product":
        raise HTTPException(status_code=400, detail="잘못된 작업 유형입니다.")
    if job.creator_user_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="접근 권한이 없습니다.")
    return StreamingResponse(
        _stream_job_events(job),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )
